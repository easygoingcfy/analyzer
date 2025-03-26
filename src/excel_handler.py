import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import os

class ExcelHandler:
    @staticmethod
    def read_stock_data(file_path, sheet_name=0):
        """从Excel文件读取股票历史数据"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            # 确保日期列为datetime类型
            if '日期' in df.columns:
                df['日期'] = pd.to_datetime(df['日期'])
            elif 'date' in df.columns:
                df['日期'] = pd.to_datetime(df['date'])
                df.rename(columns={'date': '日期'}, inplace=True)
                
            # 确保有"收盘"列
            if '收盘' not in df.columns and '收盘价' in df.columns:
                df.rename(columns={'收盘价': '收盘'}, inplace=True)
            
            return df
        except Exception as e:
            print(f"读取Excel文件时发生错误: {e}")
            return None
    
    @staticmethod
    def write_backtest_results(results_df, output_path):
        """将回测结果写入Excel文件"""
        try:
            # 处理日期格式，只保留日期部分
            if '日期' in results_df.columns:
                results_df['日期'] = pd.to_datetime(results_df['日期']).dt.date
                
            # 创建一个Excel Writer对象
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            
            # 将DataFrame写入Excel
            results_df.to_excel(writer, sheet_name='交易记录', index=False)
            
            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['交易记录']
            
            # 设置标题行格式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            
            for col in range(1, len(results_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
            
            # 设置买入单元格为绿色，卖出单元格为红色
            action_col = None
            for i, col_name in enumerate(results_df.columns):
                if col_name == '操作':
                    action_col = i + 1
                    break
            
            if action_col:
                for row in range(2, len(results_df) + 2):
                    cell = worksheet.cell(row=row, column=action_col)
                    if cell.value == '买入':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif cell.value == '卖出':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # 保存文件
            writer.close()
            print(f"回测结果已保存到: {output_path}")
            return True
        
        except Exception as e:
            print(f"写入Excel文件时发生错误: {e}")
            return False